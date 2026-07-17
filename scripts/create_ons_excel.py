#!/usr/bin/env python3
"""
Create comprehensive Excel file with all extracted data from ONS.dz
(Office National des Statistiques - Algeria)
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'skills', 'xlsx', 'templates'))
from base import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/home/z/my-project/download/ONS_dz_Complete_Data_Extraction.xlsx"

wb = Workbook()
wb.properties.creator = "Z.ai"
wb.properties.title = "ONS.dz Complete Data Extraction"
wb.properties.subject = "Algeria National Statistics Office - Full Site Data"

# ============================================================
# SHEET 1: Site Overview
# ============================================================
ws1 = wb.active
ws1.title = "Site Overview"
ws1.sheet_view.showGridLines = False

overview_data = [
    ["Website Name", "Office National des Statistiques (ONS)"],
    ["URL", "https://www.ons.dz/"],
    ["Arabic Name", "\u0627\u0644\u062f\u064a\u0648\u0627\u0646 \u0627\u0644\u0648\u0637\u0646\u064a \u0644\u0644\u0625\u062d\u0635\u0627\u0626\u064a\u0627\u062a"],
    ["Country", "Algeria"],
    ["Institution Type", "National Statistical Office (Central Statistical Institution)"],
    ["Legal Status", "\u00c9tablissement public \u00e0 caract\u00e8re administratif"],
    ["Authority", "Placed under the authority of the Prime Minister (since Nov 10, 2025, Presidential Decree N\u00b0 25-293)"],
    ["Founded", "1964 (as CNRP, renamed ONS in 1995)"],
    ["CMS", "SPIP (French open-source CMS)"],
    ["Server", "Apache/2.0.63 (Win32) PHP/5.2.9-2"],
    ["Primary Language", "French (with Arabic and English sections)"],
    ["Last Site Update", "Wednesday, July 15, 2026"],
    ["Total Visitors (displayed)", "5,839,687"],
    ["Contact Email", "ons@ons.dz ; stat@ons.dz"],
    ["Main Address", "Boulevard Mohamed Belkacemi El Annasser, Alger"],
    ["Phone", "+213 (0)23.73.81.18"],
    ["Fax", "+213 (0) 23.73.81.59"],
    ["Mission", "Collection, processing, and dissemination of socio-economic statistical information"],
    ["Key Activities", "Population & housing census, labor surveys, industrial enterprise surveys, economic accounts, price indices, trade statistics"],
]

last_col = 3
setup_sheet(ws1, title="ONS.dz - Site Overview", last_col=last_col)

headers = ["Field", "Details"]
for col_idx, header in enumerate(headers, start=2):
    ws1.cell(row=4, column=col_idx, value=header)
style_header_row(ws1, row_num=4, col_start=2, col_end=last_col)

for i, (key, val) in enumerate(overview_data):
    row_num = 5 + i
    ws1.cell(row=row_num, column=2, value=key)
    ws1.cell(row=row_num, column=3, value=val)
    style_data_row(ws1, row_num=row_num, col_start=2, col_end=last_col, row_index=i)
    ws1.cell(row=row_num, column=2).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)

ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 90

# ============================================================
# SHEET 2: Site Navigation Structure
# ============================================================
ws2 = wb.create_sheet("Navigation Structure")
ws2.sheet_view.showGridLines = False

nav_data = [
    ["1", "Accueil (Home)", "/", "Main landing page with latest publications and news"],
    ["2", "Pr\u00e9sentation", "/spip.php?rubrique38", "About ONS: history, organization, missions"],
    ["3", "Statistiques Sociales", "/spip.php?rubrique3", "Population, health, education, employment, wages, social protection"],
    ["4", "Statistiques \u00c9conomiques", "/spip.php?rubrique4", "Economic accounts, trade, agriculture, energy, transport, tourism"],
    ["5", "R\u00e9pertoires", "/spip.php?rubrique294", "National directory of economic and social agents (NIS)"],
    ["6", "Indices", "/spip.php?rubrique12", "Consumer prices, industrial production, wholesale prices, unit values"],
    ["7", "Publications", "/spip.php?rubrique2", "Statistical collections, yearbooks, bulletins, Algeria in figures"],
    ["8", "Nomenclatures", "/spip.php?rubrique24", "Activity classifications (NAA), product classifications (NPA)"],
    ["9", "Contact", "/spip.php?rubrique42", "Headquarters, regional offices, antennas across Algeria"],
]

last_col = 5
setup_sheet(ws2, title="Site Navigation Structure", last_col=last_col)

headers2 = ["#", "Menu Item", "URL Path", "Description"]
for col_idx, header in enumerate(headers2, start=2):
    ws2.cell(row=4, column=col_idx, value=header)
style_header_row(ws2, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(nav_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws2.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws2, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 5
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 65

# ============================================================
# SHEET 3: Latest Publications
# ============================================================
ws3 = wb.create_sheet("Latest Publications")
ws3.sheet_view.showGridLines = False

pubs_data = [
    ["1", "Les r\u00e9pertoires des agents \u00e9conomique et sociaux - Personnes Physiques au 31/12/2023", "New", "https://www.ons.dz/IMG/pdf/Prsphys_S2_2023.pdf"],
    ["2", "Les r\u00e9pertoires des agents \u00e9conomique et sociaux - Personnes Physiques au 31/12/2022", "", "https://www.ons.dz/IMG/pdf/Prsphys_S2_2022.pdf"],
    ["3", "Indice des prix \u00e0 la Consommation - Avril 2026", "", "https://www.ons.dz/IMG/pdf/IPC_Avril2026.pdf"],
    ["4", "Enqu\u00eate sur la situation et les perspectives dans l'industrie au 2\u00e8me trimestre 2025", "", "https://www.ons.dz/IMG/pdf/industrie2T2025.pdf"],
    ["5", "Les Comptes Nationaux Trimestriels au 2\u00e8me trimestre 2025", "", "https://www.ons.dz/IMG/pdf/CNT2T2025.pdf"],
    ["6", "Indices du commerce ext\u00e9rieur de marchandises 1er semestre 2025", "", "https://www.ons.dz/IMG/pdf/indice_commerce_exterieur_s1_2025.pdf"],
    ["7", "Indice de la Production Industrielle au 2\u00e8me trimestre 2025", "", "https://www.ons.dz/IMG/pdf/I.IPI2T2025.pdf"],
    ["8", "Indice des Prix \u00e0 la Production Industrielle au 2\u00e8me trimestre 2025", "", "https://www.ons.dz/IMG/pdf/IPPI2T2025.pdf"],
    ["9", "Les Comptes \u00c9conomiques de 2021 \u00e0 2024", "", "https://www.ons.dz/IMG/pdf/comptes_econoniques__2021_2024.pdf"],
    ["10", "Indices du commerce ext\u00e9rieur 1er trimestre 2025 (Base 2019)", "", "https://www.ons.dz/IMG/pdf/commerce_ext1T2025base2019.pdf"],
    ["11", "Indices du commerce ext\u00e9rieur - Ann\u00e9e 2024 (Base 2019)", "", "https://www.ons.dz/IMG/pdf/commerce_ext_base2019anne2024.pdf"],
    ["12", "Les Tableaux des Ressources et des Emplois 2021-2023 (Base 2001)", "", "https://www.ons.dz/IMG/pdf/TRE2021_2023.pdf"],
]

last_col = 5
setup_sheet(ws3, title="Latest Publications (Nouveaut\u00e9s)", last_col=last_col)

headers3 = ["#", "Publication Title", "Status", "PDF URL"]
for col_idx, header in enumerate(headers3, start=2):
    ws3.cell(row=4, column=col_idx, value=header)
style_header_row(ws3, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(pubs_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws3.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws3, row_num=row_num, col_start=2, col_end=last_col, row_index=i)
    if row_data[2] == "New":
        ws3.cell(row=row_num, column=4).font = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_POSITIVE)

ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 5
ws3.column_dimensions['C'].width = 70
ws3.column_dimensions['D'].width = 8
ws3.column_dimensions['E'].width = 60

# ============================================================
# SHEET 4: News
# ============================================================
ws4 = wb.create_sheet("News (Actualit\u00e9s)")
ws4.sheet_view.showGridLines = False

news_data = [
    ["15/07/2026", "Consultation pour l'acquisition des climatiseurs", "spip.php?article3148", "Consultation"],
    ["14/07/2026", "Consultation pour l'acquisition des \u00e9quipements de s\u00e9curit\u00e9", "spip.php?article3147", "Consultation"],
    ["06/07/2026", "Consultation pour l'achat d'habillement", "spip.php?article3144", "Consultation"],
    ["01/07/2026", "Consultation pour l'achat de produits d'entretien", "spip.php?article3143", "Consultation"],
    ["28/06/2026", "Consultation pour l'acquisition de pneumatiques", "spip.php?article3142", "Consultation"],
    ["24/06/2026", "Consultation relative au renouvellement des licences", "spip.php?article3141", "Consultation"],
    ["11/06/2026", "IPC enregistre une hausse de 0,4% en Avril 2026 (Alger)", "spip.php?article3129", "Statistical Release"],
    ["09/06/2026", "Consultation pour la fourniture de climatiseurs", "spip.php?article3138", "Consultation"],
    ["02/06/2026", "Consultation relative \u00e0 l'achat de papeterie", "spip.php?article3136", "Consultation"],
    ["02/06/2026", "Consultation relative \u00e0 l'achat des fournitures de bureau", "spip.php?article3135", "Consultation"],
    ["02/06/2026", "Consultation pour l'achat des produits d'entretien", "spip.php?article3134", "Consultation"],
    ["09/03/2025", "R\u00e9sultats pr\u00e9liminaires - Enqu\u00eate Activit\u00e9, Emploi et Ch\u00f4mage Oct 2024 (9,7%)", "spip.php?article3059", "Key Statistic"],
    ["13/12/2023", "Les Comptes \u00c9conomiques Base 2001", "spip.php?article2958", "Economic Data"],
]

last_col = 5
setup_sheet(ws4, title="News & Announcements (Actualit\u00e9s)", last_col=last_col)

headers4 = ["Date", "Title", "Article ID", "Category"]
for col_idx, header in enumerate(headers4, start=2):
    ws4.cell(row=4, column=col_idx, value=header)
style_header_row(ws4, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(news_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws4.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws4, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws4.column_dimensions['A'].width = 3
ws4.column_dimensions['B'].width = 14
ws4.column_dimensions['C'].width = 75
ws4.column_dimensions['D'].width = 25
ws4.column_dimensions['E'].width = 20

# ============================================================
# SHEET 5: Social Statistics
# ============================================================
ws5 = wb.create_sheet("Social Statistics")
ws5.sheet_view.showGridLines = False

social_data = [
    ["Population et D\u00e9mographie", "/spip.php?rubrique13", "Population data, RGPH census results, demographic indicators"],
    ["  Population", "/spip.php?rubrique33", "Population counts, first RGPH in 1966"],
    ["  D\u00e9mographie", "/spip.php?rubrique34", "Demographic statistics, birth/death rates, migration"],
    ["Sant\u00e9", "/spip.php?rubrique14", "Health statistics, healthcare facilities, morbidity data"],
    ["Habitat", "/spip.php?rubrique15", "Housing statistics, dwelling types, amenities"],
    ["Emploi et ch\u00f4mage", "/spip.php?rubrique16", "Employment surveys, unemployment rate (9.7% in 2024), labor force"],
    ["  D\u00e9finition et note m\u00e9thodologiques", "/spip.php?rubrique50", "Methodology definitions for employment statistics"],
    ["  Les diff\u00e9rents tableaux", "/spip.php?rubrique56", "Employment data tables"],
    ["  ENET Alg\u00e9rie 2012", "/spip.php?rubrique231", "Time-use survey results"],
    ["Salaires", "/spip.php?rubrique17", "Wage statistics, average salaries by sector"],
    ["Protection Sociale", "/spip.php?rubrique18", "Social security, pensions, social welfare data"],
    ["\u00c9ducation", "/spip.php?rubrique45", "Education statistics, enrollment rates, literacy"],
    ["D\u00e9penses de consommation des m\u00e9nages", "/spip.php?rubrique200", "Household consumption expenditure surveys"],
]

last_col = 4
setup_sheet(ws5, title="Statistiques Sociales - Sections", last_col=last_col)

headers5 = ["Section / Sub-section", "URL", "Description"]
for col_idx, header in enumerate(headers5, start=2):
    ws5.cell(row=4, column=col_idx, value=header)
style_header_row(ws5, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(social_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws5.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws5, row_num=row_num, col_start=2, col_end=last_col, row_index=i)
    if row_data[0].startswith("  "):
        ws5.cell(row=row_num, column=2).font = Font(name=FONT_NAME, size=10, italic=True, color=NEUTRAL_600)

ws5.column_dimensions['A'].width = 3
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 30
ws5.column_dimensions['D'].width = 60

# ============================================================
# SHEET 6: Economic Statistics
# ============================================================
ws6 = wb.create_sheet("Economic Statistics")
ws6.sheet_view.showGridLines = False

econ_data = [
    ["Recensement \u00c9conomique 2011", "/spip.php?rubrique166", "First Economic Census of Algeria"],
    ["  Premi\u00e8re Phase", "/spip.php?rubrique339", "First phase of economic census"],
    ["  Deuxi\u00e8me Phase", "/spip.php?rubrique340", "Second phase of economic census"],
    ["Comptes \u00e9conomiques", "/spip.php?rubrique5", "National accounts, GDP, economic aggregates"],
    ["  Quelques agr\u00e9gats 2009-2013", "/spip.php?rubrique57", "Key economic aggregates"],
    ["  Compte de production 2001-2015", "/spip.php?rubrique59", "Production account"],
    ["  Les Comptes \u00c9conomiques", "/spip.php?rubrique61", "Full economic accounts"],
    ["  Tableau \u00c9conomique d'Ensemble", "/spip.php?rubrique233", "Global economic table"],
    ["  Tableau des Entr\u00e9es-Sorties", "/spip.php?rubrique234", "Input-Output table"],
    ["Monnaie 2009-2013", "/spip.php?rubrique7", "Monetary statistics"],
    ["Indicateurs d'endettement 2009-2013", "/spip.php?rubrique58", "Debt indicators"],
    ["Commerce Ext\u00e9rieur", "/spip.php?rubrique19", "Foreign trade statistics, exports, imports"],
    ["Enqu\u00eate d'Opinion industrie et commerce", "/spip.php?rubrique202", "Business opinion surveys"],
    ["Agriculture et P\u00eache", "/spip.php?rubrique306", "Agriculture and fisheries statistics"],
    ["\u00c9nergie", "/spip.php?rubrique6", "Energy statistics"],
    ["Transports", "/spip.php?rubrique20", "Transport statistics"],
    ["Tourisme", "/spip.php?rubrique21", "Tourism statistics"],
    ["Environnement", "/spip.php?rubrique303", "Environment statistics"],
    ["Postes et T\u00e9l\u00e9communications", "/spip.php?rubrique301", "Postal and telecom statistics"],
    ["Questionnaires enqu\u00eates trimestrielles", "/spip.php?rubrique302", "Quarterly enterprise survey questionnaires"],
]

last_col = 4
setup_sheet(ws6, title="Statistiques \u00c9conomiques - Sections", last_col=last_col)

headers6 = ["Section / Sub-section", "URL", "Description"]
for col_idx, header in enumerate(headers6, start=2):
    ws6.cell(row=4, column=col_idx, value=header)
style_header_row(ws6, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(econ_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws6.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws6, row_num=row_num, col_start=2, col_end=last_col, row_index=i)
    if row_data[0].startswith("  "):
        ws6.cell(row=row_num, column=2).font = Font(name=FONT_NAME, size=10, italic=True, color=NEUTRAL_600)

ws6.column_dimensions['A'].width = 3
ws6.column_dimensions['B'].width = 50
ws6.column_dimensions['C'].width = 30
ws6.column_dimensions['D'].width = 55

# ============================================================
# SHEET 7: Indices
# ============================================================
ws7 = wb.create_sheet("Indices")
ws7.sheet_view.showGridLines = False

indices_data = [
    ["Indice des prix \u00e0 la consommation (IPC)", "Mensuelle", "Base 100 (2001)", "261 articles, 791 vari\u00e9t\u00e9s, Laspeyres", "+0.4% in April 2026", "/spip.php?rubrique26"],
    ["Indice de la production industrielle (IPI)", "Trimestrielle", "Base 100 (1989)", "Laspeyres, 1989 F.P.A.M weights", "2T 2025", "/spip.php?rubrique28"],
    ["Indice des prix de gros fruits et l\u00e9gumes", "Mensuelle (trimestrielle)", "Base 100 (1992)", "58 vari\u00e9t\u00e9s, Laspeyres", "Wholesale prices", "/spip.php?rubrique27"],
    ["Indice des Prix \u00e0 la Production Industrielle (IPPI)", "Trimestrielle", "", "Industrial producer prices", "2T 2025", "/spip.php?rubrique27"],
    ["Indice des valeurs unitaires", "Multiple", "Base 100 (2011)", "Paasche, harmonic mean weighted", "Published quarterly", "/spip.php?rubrique31"],
]

last_col = 7
setup_sheet(ws7, title="Indices (Price & Production Indices)", last_col=last_col)

headers7 = ["Index Name", "Frequency", "Base", "Methodology", "Latest Data", "URL"]
for col_idx, header in enumerate(headers7, start=2):
    ws7.cell(row=4, column=col_idx, value=header)
style_header_row(ws7, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(indices_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws7.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws7, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws7.column_dimensions['A'].width = 3
ws7.column_dimensions['B'].width = 48
ws7.column_dimensions['C'].width = 22
ws7.column_dimensions['D'].width = 16
ws7.column_dimensions['E'].width = 38
ws7.column_dimensions['F'].width = 25
ws7.column_dimensions['G'].width = 22

# ============================================================
# SHEET 8: Publications Catalog
# ============================================================
ws8 = wb.create_sheet("Publications Catalog")
ws8.sheet_view.showGridLines = False

publications_data = [
    ["Publications disponibles sur le site", "/spip.php?rubrique40", "All publications available for download"],
    ["  Indice des prix \u00e0 la consommation", "/spip.php?rubrique124", "CPI publications"],
    ["  Collections Statistiques", "/spip.php?rubrique184", "Statistical collections"],
    ["  Donn\u00e9es Statistiques", "/spip.php?rubrique126", "Statistical data tables"],
    ["  L'Alg\u00e9rie en Quelques Chiffres", "/spip.php?rubrique127", "Algeria in key figures"],
    ["  Bulletin trimestriel des statistiques", "/spip.php?rubrique267", "Quarterly statistical bulletin"],
    ["  Annuaire Statistique de l'Alg\u00e9rie", "/spip.php?rubrique289", "Statistical yearbook"],
    ["  R\u00e9trospective", "/spip.php?rubrique212", "Historical retrospective data"],
    ["  R\u00e9trospective des Comptes \u00c9conomiques", "/spip.php?rubrique375", "Economic accounts retrospective"],
    ["  R\u00e9trospective statistiques agriculture", "/spip.php?rubrique392", "Agricultural statistics retrospective"],
    ["  \u00c9ducation", "/spip.php?rubrique391", "Education statistics publications"],
    ["Publications en arabe", "/spip.php?rubrique207", "Publications in Arabic"],
    ["Publications in English", "/spip.php?rubrique220", "Publications in English"],
    ["Catalogue et modalit\u00e9s d'abonnements", "/spip.php?rubrique39", "Subscription catalog"],
    ["Publication Euro-Med", "/spip.php?rubrique173", "Euro-Med publications"],
    ["Calendrier ONS", "/spip.php?rubrique397", "Publication calendar"],
    ["RGPH 2022", "/spip.php?rubrique390", "6th Population & Housing Census (Sept 25 - Oct 16, 2022)"],
]

last_col = 4
setup_sheet(ws8, title="Publications Catalog", last_col=last_col)

headers8 = ["Publication Category", "URL", "Description"]
for col_idx, header in enumerate(headers8, start=2):
    ws8.cell(row=4, column=col_idx, value=header)
style_header_row(ws8, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(publications_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws8.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws8, row_num=row_num, col_start=2, col_end=last_col, row_index=i)
    if row_data[0].startswith("  "):
        ws8.cell(row=row_num, column=2).font = Font(name=FONT_NAME, size=10, italic=True, color=NEUTRAL_600)

ws8.column_dimensions['A'].width = 3
ws8.column_dimensions['B'].width = 50
ws8.column_dimensions['C'].width = 30
ws8.column_dimensions['D'].width = 55

# ============================================================
# SHEET 9: Repertoires & NIS
# ============================================================
ws9 = wb.create_sheet("Repertoires & NIS")
ws9.sheet_view.showGridLines = False

nis_data = [
    ["Les Personnes Morales", "/spip.php?rubrique69", "Legal entities directory"],
    ["Les Personnes Physiques", "/spip.php?rubrique80", "Natural persons (self-employed) directory"],
    ["Administrations et services d\u00e9concentr\u00e9s", "/spip.php?rubrique81", "Decentralized state services directory"],
    ["Pour tout savoir sur le NIS", "/spip.php?rubrique161", "Complete NIS system information"],
    ["Proc\u00e9dure Transitoire NIS", "/spip.php?rubrique383", "Transitional NIS attribution procedure"],
    ["D\u00e9cret l\u00e9gislatif n\u00b094-01 (PDF)", "/IMG/pdf/DN-94-01-15-01-1994.pdf", "Legal decree for statistical ID"],
    ["R\u00e9partition NIS 58 wilayas 2024 (PDF)", "/IMG/pdf/NIS_Rep58wil2024.pdf", "NIS distribution by wilaya"],
    ["Formulaire NIS PM/PP (PDF)", "/IMG/Formulaire_NIS_PM_PP.pdf", "NIS form - companies & individuals"],
    ["Formulaire NIS Administration (PDF)", "/IMG/Formulaire_NIS_ADMIN.pdf", "NIS form - government entities"],
]

last_col = 4
setup_sheet(ws9, title="R\u00e9pertoires & NIS", last_col=last_col)

headers9 = ["Section / Document", "URL / Path", "Description"]
for col_idx, header in enumerate(headers9, start=2):
    ws9.cell(row=4, column=col_idx, value=header)
style_header_row(ws9, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(nis_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws9.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws9, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws9.column_dimensions['A'].width = 3
ws9.column_dimensions['B'].width = 48
ws9.column_dimensions['C'].width = 42
ws9.column_dimensions['D'].width = 45

# ============================================================
# SHEET 10: Nomenclatures
# ============================================================
ws10 = wb.create_sheet("Nomenclatures")
ws10.sheet_view.showGridLines = False

nom_data = [
    ["Versions ant\u00e9rieures nomenclatures", "/spip.php?rubrique283", "Previous classification versions"],
    ["NAA Rev1 (Fran\u00e7ais)", "/spip.php?rubrique284", "Algerian Activities Nomenclature - French"],
    ["NPA Rev1 (Fran\u00e7ais)", "/spip.php?rubrique285", "Algerian Products Nomenclature - French"],
    ["NAA Rev1 (Arabe)", "/spip.php?rubrique381", "Algerian Activities Nomenclature - Arabic"],
    ["NPA Rev1 (Arabe)", "/spip.php?rubrique382", "Algerian Products Nomenclature - Arabic"],
    ["Nomenclatures sociales", "/spip.php?rubrique286", "Social nomenclatures"],
    ["Base de donn\u00e9es Nomenclatures", "/spip.php?rubrique287", "Nomenclatures database"],
    ["Bulletins ONU classifications", "/spip.php?rubrique288", "UN classification bulletins"],
]

last_col = 4
setup_sheet(ws10, title="Nomenclatures (Classifications)", last_col=last_col)

headers10 = ["Classification", "URL", "Description"]
for col_idx, header in enumerate(headers10, start=2):
    ws10.cell(row=4, column=col_idx, value=header)
style_header_row(ws10, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(nom_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws10.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws10, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws10.column_dimensions['A'].width = 3
ws10.column_dimensions['B'].width = 50
ws10.column_dimensions['C'].width = 30
ws10.column_dimensions['D'].width = 50

# ============================================================
# SHEET 11: Contact & Offices
# ============================================================
ws11 = wb.create_sheet("Contact & Offices")
ws11.sheet_view.showGridLines = False

contact_data = [
    ["Direction G\u00e9n\u00e9rale", "Blvd Mohamed Belkacemi El Annasser, Alger", "+213 (0)23.73.81.18", "+213 (0)23.73.81.59", "ons@ons.dz ; stat@ons.dz", "DG-DGA, DSTCNS, Inspection, DAM, DSEC, DCN, DSSR, DTIR"],
    ["Si\u00e8ge secondaire DG", "", "", "", "", ""],
    ["Annexe R\u00e9gionale Centre", "12 rue Bab Azoun, Alger", "+213 (0)20 05 20 95", "", "ara@ons.dz", ""],
    ["Annexe R\u00e9gionale Est", "Constantine", "", "", "", ""],
    ["Annexe R\u00e9gionale Ouest", "Oran", "", "", "", ""],
    ["Annexe R\u00e9gionale Sud", "Ouargla", "", "", "", ""],
    ["Antenne d'Ain D\u00e9fla", "R\u00e9gion Alger", "+213 (0)27.50.45.92", "", "", ""],
    ["Antenne de Tizi Ouzou", "R\u00e9gion Alger", "+213 (0)26.10.61.31", "", "", ""],
    ["Antenne de B\u00e9char", "R\u00e9gion Oran", "+213 (0)49.23 80 30", "", "", ""],
    ["Antenne d'Annaba", "R\u00e9gion Constantine", "+213(0)38.45.02.99", "+213(0)38.46.03.77", "", ""],
    ["Antenne de Gharda\u00efa", "R\u00e9gion Ouargla", "+213 (0)29.25.51.56", "", "", ""],
]

last_col = 7
setup_sheet(ws11, title="Contact & Regional Offices", last_col=last_col)

headers11 = ["Office", "Address", "Phone", "Fax", "Email", "Departments"]
for col_idx, header in enumerate(headers11, start=2):
    ws11.cell(row=4, column=col_idx, value=header)
style_header_row(ws11, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(contact_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws11.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws11, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws11.column_dimensions['A'].width = 3
ws11.column_dimensions['B'].width = 35
ws11.column_dimensions['C'].width = 38
ws11.column_dimensions['D'].width = 22
ws11.column_dimensions['E'].width = 22
ws11.column_dimensions['F'].width = 28
ws11.column_dimensions['G'].width = 48

# ============================================================
# SHEET 12: All PDF Documents
# ============================================================
ws12 = wb.create_sheet("All PDF Documents")
ws12.sheet_view.showGridLines = False

pdf_data = [
    ["Homepage", "R\u00e9pertoires Personnes Physiques S2 2023", "https://www.ons.dz/IMG/pdf/Prsphys_S2_2023.pdf"],
    ["Homepage", "R\u00e9pertoires Personnes Physiques S2 2022", "https://www.ons.dz/IMG/pdf/Prsphys_S2_2022.pdf"],
    ["Homepage", "IPC Avril 2026", "https://www.ons.dz/IMG/pdf/IPC_Avril2026.pdf"],
    ["Homepage", "Enqu\u00eate Industrie 2T 2025", "https://www.ons.dz/IMG/pdf/industrie2T2025.pdf"],
    ["Homepage", "Comptes Nationaux Trimestriels 2T 2025", "https://www.ons.dz/IMG/pdf/CNT2T2025.pdf"],
    ["Homepage", "Indices Commerce Ext\u00e9rieur S1 2025", "https://www.ons.dz/IMG/pdf/indice_commerce_exterieur_s1_2025.pdf"],
    ["Homepage", "IPI 2T 2025", "https://www.ons.dz/IMG/pdf/I.IPI2T2025.pdf"],
    ["Homepage", "IPPI 2T 2025", "https://www.ons.dz/IMG/pdf/IPPI2T2025.pdf"],
    ["Homepage", "Comptes \u00c9conomiques 2021-2024", "https://www.ons.dz/IMG/pdf/comptes_econoniques__2021_2024.pdf"],
    ["Homepage", "Commerce Ext\u00e9rieur 1T 2025 (Base 2019)", "https://www.ons.dz/IMG/pdf/commerce_ext1T2025base2019.pdf"],
    ["Homepage", "Commerce Ext\u00e9rieur Ann\u00e9e 2024 (Base 2019)", "https://www.ons.dz/IMG/pdf/commerce_ext_base2019anne2024.pdf"],
    ["Homepage", "TRE 2021-2023 (Base 2001)", "https://www.ons.dz/IMG/pdf/TRE2021_2023.pdf"],
    ["Sidebar", "Calendrier ONS 2025", "https://www.ons.dz/IMG/pdf/CalendrierONS_2025.pdf"],
    ["NIS", "D\u00e9cret l\u00e9gislatif n\u00b094-01", "https://www.ons.dz/IMG/pdf/DN-94-01-15-01-1994.pdf"],
    ["NIS", "R\u00e9partition NIS 58 wilayas 2024", "https://www.ons.dz/IMG/pdf/NIS_Rep58wil2024.pdf"],
    ["IPC", "M\u00e9thodologie IPC", "https://www.ons.dz/IMG/pdf/ipc_meth.pdf"],
    ["IPC", "IPC Novembre 2019", "https://www.ons.dz/IMG/pdf/ipc_novembre2019-2.pdf"],
    ["IPC", "\u00c9volution par groupe Nov 2019", "https://www.ons.dz/IMG/pdf/evolution_group11_2019.pdf"],
    ["Commerce Ext.", "10 premiers clients export 92-06", "https://www.ons.dz/IMG/pdf/10_1er_clients_expo_92-06.pdf"],
    ["Commerce Ext.", "\u00c9volution exportations par pays 92-06", "https://www.ons.dz/IMG/pdf/Evol_expo_pays_de_destination92-06.pdf"],
    ["Commerce Ext.", "\u00c9volution exportations par secteur", "https://www.ons.dz/IMG/pdf/Evol_expo_secteur_active.pdf"],
    ["Commerce Ext.", "\u00c9volution exportations par branche NAPR", "https://www.ons.dz/IMG/pdf/Evol_expo_N.A.P.R.pdf"],
    ["Commerce Ext.", "\u00c9volution CGCE 1996-2006", "https://www.ons.dz/IMG/pdf/Evol_C.G.C.E_96-06.pdf"],
    ["Commerce Ext.", "20 premiers clients hors hydro. 92-06", "https://www.ons.dz/IMG/pdf/20_1er_clients_92-06.pdf"],
    ["Commerce Ext.", "Exportations hors hydrocarbures 2006", "https://www.ons.dz/IMG/pdf/Expo_prod_hydrocarbure2006.pdf"],
    ["Commerce Ext.", "10 premiers pays fournisseurs 92-06", "https://www.ons.dz/IMG/pdf/10_1er_pays_fournisseur92-06.pdf"],
    ["Commerce Ext.", "Importations par pays d'origine 92-06", "https://www.ons.dz/IMG/pdf/imp_march_pays_origine92-06.pdf"],
    ["Commerce Ext.", "Importations par NAA", "https://www.ons.dz/IMG/pdf/Evol_imp_nomenclature.pdf"],
    ["Liens utiles", "Liste instituts nationaux statistique", "http://www.ons.dz/IMG/Liste_des_instituts_nationaux_de_statistique.pdf"],
    ["Collections", "Population par \u00e2ge, sexe et commune", "https://www.ons.dz/collections/w42_p3.pdf"],
    ["Collections", "Population par commune et dispersion (w26)", "https://www.ons.dz/collections/w26_p2.pdf"],
    ["Collections", "Structure par \u00e2ge et sexe population", "https://www.ons.dz/collections/w06_p2.pdf"],
    ["Collections", "Population par commune (w13)", "https://www.ons.dz/collections/w13_p2.pdf"],
]

last_col = 4
setup_sheet(ws12, title="All PDF Documents Found on Site", last_col=last_col)

headers12 = ["Source", "Document Title", "Full URL"]
for col_idx, header in enumerate(headers12, start=2):
    ws12.cell(row=4, column=col_idx, value=header)
style_header_row(ws12, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(pdf_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws12.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws12, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws12.column_dimensions['A'].width = 3
ws12.column_dimensions['B'].width = 15
ws12.column_dimensions['C'].width = 48
ws12.column_dimensions['D'].width = 70

# ============================================================
# SHEET 13: Key Statistics
# ============================================================
ws13 = wb.create_sheet("Key Statistics")
ws13.sheet_view.showGridLines = False

stats_data = [
    ["Unemployment Rate (2024)", "9.7%", "Enqu\u00eate Activit\u00e9, Emploi et Ch\u00f4mage Oct 2024", "spip.php?article3059"],
    ["IPC Variation (Apr 2026 vs Mar 2026)", "+0.4%", "IPC - Ville d'Alger", "spip.php?article3129"],
    ["IPC Variation (Apr 2025 vs Mar 2025)", "-0.7%", "IPC - Ville d'Alger", "spip.php?article3129"],
    ["Total Site Visitors", "5,839,687", "Homepage footer", "https://www.ons.dz/"],
    ["RGPH 2022 Census Period", "25 Sept - 16 Oct 2022", "6th General Population & Housing Census", "spip.php?rubrique390"],
    ["ONS Founded", "1964", "As CNRP", "spip.php?rubrique277"],
    ["First Census (Independent Algeria)", "1966", "Conducted by CNRP", "spip.php?rubrique33"],
    ["NIS Code Length", "15 digits", "Num\u00e9ro d'Identification Statistique", "spip.php?rubrique161"],
    ["IPC Coverage", "261 articles, 791 varieties", "Base 2001", "spip.php?rubrique26"],
    ["Regional Annexes", "4 (Centre, Est, Ouest, Sud)", "Covering all Algeria", "spip.php?rubrique42"],
    ["Regional Antennas", "5", "Ain Defla, Tizi Ouzou, B\u00e9char, Annaba, Gharda\u00efa", "spip.php?rubrique402"],
]

last_col = 5
setup_sheet(ws13, title="Key Statistics Found on Site", last_col=last_col)

headers13 = ["Statistic", "Value", "Context", "Source"]
for col_idx, header in enumerate(headers13, start=2):
    ws13.cell(row=4, column=col_idx, value=header)
style_header_row(ws13, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(stats_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws13.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws13, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws13.column_dimensions['A'].width = 3
ws13.column_dimensions['B'].width = 42
ws13.column_dimensions['C'].width = 30
ws13.column_dimensions['D'].width = 48
ws13.column_dimensions['E'].width = 30

# ============================================================
# SHEET 14: All URLs
# ============================================================
ws14 = wb.create_sheet("All URLs")
ws14.sheet_view.showGridLines = False

all_urls = [
    ["Main", "Homepage", "https://www.ons.dz/", "200"],
    ["Main", "Pr\u00e9sentation", "https://www.ons.dz/spip.php?rubrique38", "200"],
    ["Main", "Statistiques Sociales", "https://www.ons.dz/spip.php?rubrique3", "200"],
    ["Main", "Statistiques \u00c9conomiques", "https://www.ons.dz/spip.php?rubrique4", "200"],
    ["Main", "R\u00e9pertoires", "https://www.ons.dz/spip.php?rubrique294", "200"],
    ["Main", "Indices", "https://www.ons.dz/spip.php?rubrique12", "200"],
    ["Main", "Publications", "https://www.ons.dz/spip.php?rubrique2", "200"],
    ["Main", "Nomenclatures", "https://www.ons.dz/spip.php?rubrique24", "200"],
    ["Main", "Contact", "https://www.ons.dz/spip.php?rubrique42", "200"],
    ["Presentation", "Historique", "https://www.ons.dz/spip.php?rubrique277", "200"],
    ["Presentation", "Syst\u00e8me National Statistique", "https://www.ons.dz/spip.php?rubrique278", "200"],
    ["Presentation", "Conseil National Statistique", "https://www.ons.dz/spip.php?rubrique279", "200"],
    ["Presentation", "Fonctions de l'ONS", "https://www.ons.dz/spip.php?rubrique280", "200"],
    ["Presentation", "Organisation interne", "https://www.ons.dz/spip.php?rubrique281", "200"],
    ["Presentation", "Organigramme", "https://www.ons.dz/spip.php?rubrique282", "200"],
    ["Social", "Population et D\u00e9mographie", "https://www.ons.dz/spip.php?rubrique13", "200"],
    ["Social", "Sant\u00e9", "https://www.ons.dz/spip.php?rubrique14", "200"],
    ["Social", "Habitat", "https://www.ons.dz/spip.php?rubrique15", "200"],
    ["Social", "Emploi et ch\u00f4mage", "https://www.ons.dz/spip.php?rubrique16", "200"],
    ["Social", "Salaires", "https://www.ons.dz/spip.php?rubrique17", "200"],
    ["Social", "Protection Sociale", "https://www.ons.dz/spip.php?rubrique18", "200"],
    ["Social", "\u00c9ducation", "https://www.ons.dz/spip.php?rubrique45", "200"],
    ["Social", "D\u00e9penses consommation m\u00e9nages", "https://www.ons.dz/spip.php?rubrique200", "200"],
    ["Economic", "Comptes \u00e9conomiques", "https://www.ons.dz/spip.php?rubrique5", "200"],
    ["Economic", "Commerce Ext\u00e9rieur", "https://www.ons.dz/spip.php?rubrique19", "200"],
    ["Economic", "\u00c9nergie", "https://www.ons.dz/spip.php?rubrique6", "200"],
    ["Economic", "Transports", "https://www.ons.dz/spip.php?rubrique20", "200"],
    ["Economic", "Tourisme", "https://www.ons.dz/spip.php?rubrique21", "200"],
    ["Economic", "Recensement \u00c9conomique 2011", "https://www.ons.dz/spip.php?rubrique166", "200"],
    ["Special", "Collections (frame-based)", "https://www.ons.dz/collections", "WAF Blocked"],
    ["Special", "RGPH 2020", "https://www.ons.dz/rgph2020", "500 Error"],
    ["Special", "RGPH 2022", "https://www.ons.dz/spip.php?rubrique390", "200"],
    ["Special", "Actualit\u00e9s archive", "https://www.ons.dz/spip.php?rubrique226", "200"],
    ["Special", "Liens utiles", "https://www.ons.dz/spip.php?rubrique139", "200"],
    ["Special", "Webmail ONS", "https://webmail.ons.dz/owa", "External"],
]

last_col = 5
setup_sheet(ws14, title="Complete URL Inventory", last_col=last_col)

headers14 = ["Category", "Page Name", "Full URL", "Status"]
for col_idx, header in enumerate(headers14, start=2):
    ws14.cell(row=4, column=col_idx, value=header)
style_header_row(ws14, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(all_urls):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws14.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws14, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws14.column_dimensions['A'].width = 3
ws14.column_dimensions['B'].width = 15
ws14.column_dimensions['C'].width = 42
ws14.column_dimensions['D'].width = 55
ws14.column_dimensions['E'].width = 12

# ============================================================
# SHEET 15: ONS History
# ============================================================
ws15 = wb.create_sheet("ONS History")
ws15.sheet_view.showGridLines = False

history_data = [
    ["1964", "Creation of CNRP", "Commissariat National pour le Recensement de la Population, created after independence"],
    ["1966", "First Population Census (RGPH)", "First census of independent Algeria, conducted by CNRP"],
    ["1971", "Renamed to CNRES", "Commissariat National aux Recensements et Enqu\u00eates Statistiques"],
    ["1972-1973", "Demographic Survey", "First national demographic survey"],
    ["1972-1975", "Cartographic Survey", "National cartographic mapping survey"],
    ["1977", "Second RGPH", "Second General Population and Housing Census"],
    ["1979-1980", "Household Consumption Survey", "Enqu\u00eate sur la consommation des m\u00e9nages"],
    ["1982", "Legislative Decree N\u00b0 82-489", "Reorganization of the national statistics system"],
    ["1985", "Decree N\u00b0 85-311", "Modification of 1982 decree"],
    ["1987", "Third RGPH", "Third General Population and Housing Census"],
    ["1994", "Legislative Decree N\u00b0 94-01", "Established the national directory and NIS system"],
    ["1995", "Decree N\u00b0 95-159", "Current ONS structure established"],
    ["2011", "First Economic Census", "Premier Recensement \u00c9conomique de l'Alg\u00e9rie"],
    ["2012", "Time-Use Survey (ENET)", "Enqu\u00eate Nationale sur l'Emploi du Temps"],
    ["2022", "6th RGPH", "Sept 25 - Oct 16, 2022"],
    ["2025", "Presidential Decree N\u00b0 25-293", "ONS placed under the Prime Minister's authority (Nov 10)"],
]

last_col = 4
setup_sheet(ws15, title="ONS History (Historique)", last_col=last_col)

headers15 = ["Year", "Event", "Description"]
for col_idx, header in enumerate(headers15, start=2):
    ws15.cell(row=4, column=col_idx, value=header)
style_header_row(ws15, row_num=4, col_start=2, col_end=last_col)

for i, row_data in enumerate(history_data):
    row_num = 5 + i
    for col_idx, value in enumerate(row_data, start=2):
        ws15.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws15, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

ws15.column_dimensions['A'].width = 3
ws15.column_dimensions['B'].width = 12
ws15.column_dimensions['C'].width = 40
ws15.column_dimensions['D'].width = 70

# ============================================================
# Save
# ============================================================
wb.save(OUTPUT_PATH)
print(f"Excel file saved to: {OUTPUT_PATH}")
print(f"Sheets: {wb.sheetnames}")